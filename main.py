# main.py - PostgreSQL Only Version

import os
import sys
import io
import csv
from flask import Flask, render_template, request, redirect, url_for, flash, session, Response, jsonify, send_from_directory, abort
from datetime import datetime, timedelta, date
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import secrets
import re
import psycopg
from psycopg.rows import dict_row
from psycopg.errors import UniqueViolation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Carregar variáveis de ambiente do arquivo .env se existir
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv não está instalado, usar apenas variáveis de ambiente do sistema

# Inicializa a aplicação Flask
app = Flask(__name__)
# Usa SECRET_KEY do ambiente em produção; gera uma chave temporária caso não definida
app.secret_key = os.getenv('SECRET_KEY', secrets.token_hex(16))  # Chave secreta para sessões

# --- Configuração SQLAlchemy ---
from app.models import (
    db, User, Client, Customer, Loan, Document,
    Payment, PaymentHistory, Notification, Installment, Configuration
)
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, case, or_
from flask_migrate import Migrate

DATABASE_URL = os.getenv('DATABASE_URL')
if not DATABASE_URL:
    raise RuntimeError('DATABASE_URL não configurada. Defina a variável de ambiente.')

db_url = DATABASE_URL
if db_url.startswith("postgresql://"):
    db_url = db_url.replace("postgresql://", "postgresql+psycopg://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# --- Fim da Configuração ---

# --- Inicialização das Extensões ---
db.init_app(app)
migrate = Migrate(app, db)
# --- Fim da Inicialização ---

# Configuração de uploads
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


# --- Configuração e Inicialização do Banco de Dados ---

def get_db():
    """Abre uma nova conexão com o banco de dados PostgreSQL."""
    # Em ambientes gerenciados (ex.: Render), forçar SSL se não especificado
    db_url = DATABASE_URL
    if 'sslmode=' not in db_url and 'localhost' not in db_url and '127.0.0.1' not in db_url:
        separator = '&' if '?' in db_url else '?'
        db_url = f"{db_url}{separator}sslmode=require"
    
    conn = psycopg.connect(db_url)
    return conn

# --- Decoradores de Autenticação ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        if session.get('usuario_tipo') != 'admin':
            flash('Acesso negado. Apenas administradores podem acessar esta página.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def gerente_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'critical')
            return redirect(url_for('login'))
        
        # Buscar nível do usuário no banco de dados
        usuario = User.query.get(session['usuario_id'])
        
        if not usuario or usuario.nivel not in ['Gerente', 'ADM']:
            flash('Acesso não autorizado.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def adm_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        
        # Buscar nível do usuário no banco de dados
        usuario = User.query.get(session['usuario_id'])
        
        if not usuario or usuario.nivel != 'ADM':
            flash('Acesso não autorizado. Apenas administradores.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# --- Funções Auxiliares ---
def calcular_valor_atualizado(cobranca):
    """Calcula o valor atualizado de uma cobrança com juros e multa."""
    if cobranca['status'] != 'Pago' and cobranca['data_vencimento']:
        # Handle both string and date object formats
        if isinstance(cobranca['data_vencimento'], str):
            data_venc = datetime.strptime(cobranca['data_vencimento'], '%Y-%m-%d')
        else:
            # It's already a date object, convert to datetime
            data_venc = datetime.combine(cobranca['data_vencimento'], datetime.min.time())
        hoje = datetime.now()
        
        if hoje > data_venc:
            dias_atraso = (hoje - data_venc).days
            
            # Buscar configurações
            conn = get_db()
            cur = conn.cursor(row_factory=dict_row)
            cur.execute('SELECT chave, valor FROM configuracoes')
            config_rows = cur.fetchall()
            config = {row['chave']: row['valor'] for row in config_rows}
            cur.close()
            conn.close()
            
            dias_tolerancia = int(config.get('dias_tolerancia', 3))
            
            if dias_atraso > dias_tolerancia:
                # Aplicar multa
                taxa_multa = float(config.get('taxa_multa', 10.0))
                multa = cobranca['valor_original'] * (taxa_multa / 100)
                
                # Aplicar juros
                taxa_juros = float(config.get('taxa_juros_mensal', 2.0))
                meses_atraso = dias_atraso / 30
                juros = cobranca['valor_original'] * (taxa_juros / 100) * meses_atraso
                
                return {
                    'multa': round(multa, 2),
                    'juros': round(juros, 2),
                    'valor_total': round(cobranca['valor_original'] + multa + juros - cobranca.get('desconto', 0), 2),
                    'dias_atraso': dias_atraso
                }
    
    return {
        'multa': 0,
        'juros': 0,
        'valor_total': cobranca['valor_original'] - cobranca.get('desconto', 0),
        'dias_atraso': 0
    }

def validar_cpf_cnpj(documento):
    """Valida CPF ou CNPJ."""
    # Remove caracteres não numéricos
    documento = re.sub(r'\D', '', documento)
    
    if len(documento) == 11:  # CPF
        # Validação simplificada do CPF
        return len(documento) == 11
    elif len(documento) == 14:  # CNPJ
        # Validação simplificada do CNPJ
        return len(documento) == 14
    
    return False

# --- Rotas de Autenticação ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        
        usuario = User.query.filter_by(email=email).first()
        
        if usuario and check_password_hash(usuario.senha, senha):
            session['usuario_id'] = usuario.id
            session['usuario_nome'] = usuario.nome
            # Derive tipo from nivel
            nivel = usuario.nivel or 'Operador'
            tipo = 'admin' if nivel == 'ADM' else 'operador'
            session['usuario_tipo'] = tipo
            session['usuario_nivel'] = nivel
            flash(f'Bem-vindo, {usuario.nome}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Email ou senha incorretos.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Você foi desconectado com sucesso.', 'info')
    return redirect(url_for('login'))

# --- Rotas do Dashboard ---
@app.route('/')
@login_required
def index():
    """Renderiza o dashboard com estatísticas e lista de cobranças."""
    
    # --- 1. Estatísticas Gerais (Counts) ---
    total_clientes = Customer.query.count()
    cobrancas_pendentes = Loan.query.filter_by(status='Pendente').count()
    
    # Cobranças Vencidas (Status Pendente e Data Vencimento < Hoje)
    # Nota: A data de vencimento da COBRANÇA é a da última parcela ou a inicial, 
    # mas aqui estamos contando quantas *cobranças* (contratos) estão tecnicamente atrasadas
    cobrancas_vencidas = Loan.query.filter(
        Loan.status == 'Pendente', 
        Loan.data_vencimento < date.today()
    ).count()
    
    cobrancas_pagas = Loan.query.filter_by(status='Pago').count()
    
    # --- 2. Cálculo Financeiro (Saldo Devedor Total) ---
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    
    # Buscar todas as parcelas pendentes de cobranças ativas
    # Isso é mais eficiente do que buscar cobranças e depois parcelas
    parcelas_pendentes_query = db.session.query(Installment).join(Loan).filter(
        Loan.status == 'Pendente',
        Installment.status == 'Pendente'
    ).all()
    
    saldo_devedor_total = 0
    for parcela in parcelas_pendentes_query:
        valor_base = float(parcela.valor)
        multa = float(parcela.multa_manual or 0)
        saldo_devedor_total += (valor_base + multa)
    
    # Total recebido no mês (Soma da tabela Payments)
    total_recebido_mes = db.session.query(func.sum(Payment.valor_pago))\
        .filter(Payment.data_pagamento >= primeiro_dia_mes).scalar() or 0
    
    # --- 3. KPIs por Empresa (FH1, FH2...) ---
    empresas = ['FH1', 'FH2', 'FH3', 'FH4']
    kpis_por_empresa = {}
    
    for emp in empresas:
        # Clientes por empresa
        count_cli = Customer.query.filter_by(empresa=emp).count()
        
        # Saldo devedor por empresa (Complexo, exige join)
        # Soma valor das parcelas pendentes onde o cliente é da empresa X
        saldo_emp = db.session.query(func.sum(Installment.valor))\
            .join(Loan).join(Customer)\
            .filter(
                Customer.empresa == emp,
                Loan.status == 'Pendente',
                Installment.status == 'Pendente'
            ).scalar() or 0
            
        # Adiciona multas manuais ao saldo da empresa
        multas_emp = db.session.query(func.sum(Installment.multa_manual))\
            .join(Loan).join(Customer)\
            .filter(
                Customer.empresa == emp,
                Loan.status == 'Pendente',
                Installment.status == 'Pendente'
            ).scalar() or 0
            
        kpis_por_empresa[emp] = {
            'total_clientes': count_cli,
            'saldo_devedor': float(saldo_emp) + float(multas_emp)
        }
    
    stats = {
        'total_clientes': total_clientes,
        'cobrancas_pendentes': cobrancas_pendentes,
        'cobrancas_vencidas': cobrancas_vencidas,
        'cobrancas_pagas': cobrancas_pagas,
        'saldo_devedor_total': saldo_devedor_total,
        'total_recebido_mes': float(total_recebido_mes),
        'kpis_por_empresa': kpis_por_empresa,
    }
    
    # --- 4. Cobranças Recentes (Listagem) ---
    recentes = Loan.query.filter(Loan.status != 'Pago')\
        .order_by(Loan.data_vencimento.asc()).limit(20).all()
    
    cobrancas_atualizadas = []
    for cob in recentes:
        # Precisamos converter para dicionário ou objeto compatível para o template
        # Vamos calcular saldo dinâmico desta cobrança
        saldo_cob = sum(
            float(p.valor) + float(p.multa_manual or 0) 
            for p in cob.installments if p.status == 'Pendente'
        )
        
        # Objeto auxiliar para o template
        cobrancas_atualizadas.append({
            'id': cob.id,
            'cliente_nome': cob.customer.nome,
            'telefone': cob.customer.telefone,
            'email': cob.customer.email,
            'descricao': cob.descricao,
            'valor_original': cob.valor_original,
            'data_vencimento': cob.data_vencimento,
            'status': cob.status,
            'valor_pago': cob.valor_pago,
            'total_a_pagar': saldo_cob, # Substitui lógica antiga
            'saldo_devedor': saldo_cob,
            'dias_atraso': (hoje - cob.data_vencimento).days if cob.data_vencimento and hoje > cob.data_vencimento else 0
        })
    
    # --- 5. Clientes Inadimplentes (Atrasados) ---
    # Busca clientes com parcelas vencidas e não pagas
    clientes_inadimplentes = Customer.query.join(Loan).join(Installment).filter(
        Installment.data_vencimento < hoje,
        Installment.status == 'Pendente'
    ).distinct().order_by(Customer.nome).all()
    
    # Calcular saldo total de cada inadimplente
    for cli in clientes_inadimplentes:
        saldo = 0
        for cob in cli.loans:
            if cob.status == 'Pendente':
                for p in cob.installments:
                    if p.status == 'Pendente':
                        saldo += float(p.valor) + float(p.multa_manual or 0)
        cli.saldo_devedor_total = saldo

    return render_template('index.html', 
                         stats=stats, 
                         cobrancas=cobrancas_atualizadas,
                         clientes_inadimplentes=clientes_inadimplentes,
                         usuario=session)

# --- Rotas de Clientes ---
@app.route('/clientes')
@login_required
def listar_clientes():
    """Lista clientes com filtros por status e empresa."""
    hoje = date.today()

    filtro_status = request.args.get('status', 'todos')
    filtro_empresa = request.args.get('empresa')

    query = db.session.query(Customer)
    
    if filtro_empresa:
        query = query.filter(Customer.empresa == filtro_empresa)

    if filtro_status == 'atrasado':
        query = query.join(Customer.loans).join(Loan.installments).filter(
            Installment.data_vencimento < hoje,
            Installment.status == 'Pendente'
        ).distinct()

    clientes = query.order_by(Customer.nome.asc()).all()

    # Calcular saldo devedor usando relacionamentos do SQLAlchemy
    for cliente in clientes:
        saldo_devedor_cliente = 0
        cobrancas_abertas = [c for c in cliente.loans if c.status == 'Pendente']
        for cobranca in cobrancas_abertas:
            parcelas_pendentes = [p for p in cobranca.installments if p.status == 'Pendente']
            for parcela in parcelas_pendentes:
                valor_parcela_atualizado = float(parcela.valor) + (float(parcela.multa_manual) if parcela.multa_manual else 0)
                saldo_devedor_cliente += valor_parcela_atualizado
        cliente.saldo_devedor_total = saldo_devedor_cliente

    return render_template('clientes.html', clientes=clientes, 
                           filtro_ativo=filtro_status,
                           filtro_empresa=filtro_empresa)



@app.route('/cliente/adicionar', methods=['GET', 'POST'])
@login_required
def adicionar_cliente():
    """Adiciona um novo cliente."""
    if request.method == 'POST':
        files = request.files.getlist('documentos')

        dados = {
            'nome': request.form['nome'],
            'cpf_cnpj': request.form.get('cpf_cnpj', ''),
            'rg': request.form.get('rg', ''),
            'email': request.form.get('email', ''),
            'telefone': request.form['telefone'],
            'telefone_secundario': request.form.get('telefone_secundario', ''),
            'chave_pix': request.form.get('chave_pix', ''),
            'endereco': request.form.get('endereco', ''),
            'cidade': request.form.get('cidade', ''),
            'estado': request.form.get('estado', ''),
            'cep': request.form.get('cep', ''),
            'referencia': request.form.get('referencia', ''),
            'telefone_referencia': request.form.get('telefone_referencia', ''),
            'endereco_referencia': request.form.get('endereco_referencia', ''),
            'observacoes': request.form.get('observacoes', ''),
            'empresa': request.form.get('empresa', 'FH1')
        }
        
        # Validação
        if not dados['cpf_cnpj']:
            flash('CPF/CNPJ é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['rg']:
            flash('RG é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['chave_pix']:
            flash('Chave Pix é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['referencia']:
            flash('Nome da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['telefone_referencia']:
            flash('Telefone da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['endereco_referencia']:
            flash('Endereço da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['endereco']:
            flash('Endereço é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['cidade']:
            flash('Cidade é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['estado']:
            flash('Estado é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['cep']:
            flash('CEP é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if dados['cpf_cnpj'] and not validar_cpf_cnpj(dados['cpf_cnpj']):
            flash('CPF/CNPJ inválido.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['empresa']:
            flash('Empresa é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        novo_cliente = Customer(
            nome=dados['nome'],
            cpf_cnpj=dados['cpf_cnpj'],
            rg=dados['rg'],
            email=dados['email'],
            telefone=dados['telefone'],
            telefone_secundario=dados['telefone_secundario'],
            chave_pix=dados['chave_pix'],
            endereco=dados['endereco'],
            cidade=dados['cidade'],
            estado=dados['estado'],
            cep=dados['cep'],
            referencia=dados['referencia'],
            telefone_referencia=dados['telefone_referencia'],
            endereco_referencia=dados['endereco_referencia'],
            observacoes=dados['observacoes'],
            empresa=dados['empresa']
        )
        db.session.add(novo_cliente)
        
        try:
            db.session.commit()
            
            # Pega o ID após o commit para usar na pasta de upload
            novo_cliente_id = novo_cliente.id
            
            # Processar uploads de documentos
            if files:
                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        client_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(novo_cliente_id))
                        os.makedirs(client_upload_folder, exist_ok=True)
                        file_path = os.path.join(client_upload_folder, filename)
                        file.save(file_path)
                        # Manter a query SQL para documentos por enquanto
                        conn = get_db()
                        cur = conn.cursor()
                        cur.execute('INSERT INTO documentos (cliente_id, nome_ficheiro) VALUES (%s, %s)', (novo_cliente_id, filename))
                        conn.commit()
                        cur.close()
                        conn.close()
            
            flash('Cliente adicionado com sucesso!', 'success')
            return redirect(url_for('listar_clientes'))
            
        except IntegrityError:
            db.session.rollback()
            flash('CPF/CNPJ já cadastrado.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
    
    return render_template('cliente_form.html', cliente=None)

@app.route('/cliente/<int:cliente_id>')
@login_required
def visualizar_cliente(cliente_id):
    """Visualiza detalhes de um cliente específico."""
    cliente = Customer.query.get_or_404(cliente_id)
    
    # Cobranças do cliente usando relacionamento
    cobrancas = cliente.loans.order_by(Loan.data_vencimento.desc()).all()
    
    # Histórico de pagamentos com join para pegar a descrição da cobrança
    historico_query = db.session.query(
        PaymentHistory,
        Loan.descricao.label('cobranca_descricao')
    ).join(Loan).filter(
        PaymentHistory.cliente_id == cliente_id
    ).order_by(PaymentHistory.data_pagamento.desc()).all()
    
    # Adicionar atributo cobranca_descricao aos objetos PaymentHistory
    historico = []
    for pagamento, descricao in historico_query:
        # Adicionar o atributo dinamicamente ao objeto
        pagamento.cobranca_descricao = descricao
        historico.append(pagamento)
    
    # Documentos do cliente usando relacionamento
    documentos = cliente.documents.order_by(Document.id.desc()).all()

    # Nova lógica de cálculo por parcela
    hoje = date.today()
    cobrancas_processadas = []
    if cobrancas:  # Garante que só vai iterar se houver cobranças
        for cobranca in cobrancas:
            # Criar um dicionário para compatibilidade com o template
            cobranca_dict = {
                'id': cobranca.id,
                'cliente_id': cobranca.cliente_id,
                'descricao': cobranca.descricao,
                'valor_original': float(cobranca.valor_original) if cobranca.valor_original else 0,
                'valor_pago': float(cobranca.valor_pago) if cobranca.valor_pago else 0,
                'valor_total': float(cobranca.valor_total) if cobranca.valor_total else 0,
                'taxa_juros': float(cobranca.taxa_juros) if cobranca.taxa_juros else 0,
                'data_vencimento': cobranca.data_vencimento,
                'data_pagamento': cobranca.data_pagamento,
                'status': cobranca.status,
                'numero_parcelas': cobranca.numero_parcelas,
                'tipo_cobranca': cobranca.tipo_cobranca,
                'criado_em': cobranca.criado_em,
                'atualizado_em': cobranca.atualizado_em
            }
            saldo_devedor_cobranca = 0
            cobranca_dict['parcelas_com_multa'] = []  # Lista para guardar dados das parcelas para o template

            # Buscar parcelas usando relacionamento do SQLAlchemy
            parcelas = sorted(cobranca.installments, key=lambda p: p.numero_parcela)

            for parcela in parcelas:
                multa_manual = float(parcela.multa_manual) if parcela.multa_manual else 0
                valor_atualizado = float(parcela.valor) + multa_manual

                if parcela.status == 'Pendente':
                    saldo_devedor_cobranca += valor_atualizado

                # Guarda os dados calculados para exibir no template
                cobranca_dict['parcelas_com_multa'].append({
                    'numero': parcela.numero_parcela,
                    'vencimento': parcela.data_vencimento,
                    'valor_original': float(parcela.valor),
                    'multa_manual': multa_manual,
                    'valor_atualizado': valor_atualizado,
                    'status': parcela.status,
                    'id': parcela.id
                })

            # Adiciona atributos dinâmicos à cobrança para exibição consistente
            cobranca_dict['multa_aplicada'] = 0  # Multas agora são por parcela
            cobranca_dict['saldo_devedor_calculado'] = saldo_devedor_cobranca - (cobranca_dict['valor_pago'] or 0)
            
            # Para compatibilidade com template existente
            cobranca_dict['valor_multa'] = 0  # Multas agora são por parcela
            cobranca_dict['total_a_pagar'] = cobranca_dict['saldo_devedor_calculado']
            cobranca_dict['saldo_devedor'] = cobranca_dict['saldo_devedor_calculado']
            
            cobrancas_processadas.append(cobranca_dict)
    
    return render_template('cliente_detalhes.html', 
                         cliente=cliente, 
                         cobrancas=cobrancas_processadas,
                         historico=historico,
                         documentos=documentos,
                         today=hoje)

@app.route('/cliente/<int:cliente_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    """Edita um cliente existente."""
    cliente = Customer.query.get_or_404(cliente_id)
    
    if request.method == 'POST':
        dados = {
            'nome': request.form['nome'],
            'cpf_cnpj': request.form.get('cpf_cnpj', ''),
            'rg': request.form.get('rg', ''),
            'email': request.form.get('email', ''),
            'telefone': request.form['telefone'],
            'telefone_secundario': request.form.get('telefone_secundario', ''),
            'chave_pix': request.form.get('chave_pix', ''),
            'endereco': request.form.get('endereco', ''),
            'cidade': request.form.get('cidade', ''),
            'estado': request.form.get('estado', ''),
            'cep': request.form.get('cep', ''),
            'referencia': request.form.get('referencia', ''),
            'telefone_referencia': request.form.get('telefone_referencia', ''),
            'endereco_referencia': request.form.get('endereco_referencia', ''),
            'observacoes': request.form.get('observacoes', ''),
            'empresa': request.form.get('empresa', 'FH1')
        }
        
        # Validação
        if not dados['cpf_cnpj']:
            flash('CPF/CNPJ é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['rg']:
            flash('RG é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['chave_pix']:
            flash('Chave Pix é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['referencia']:
            flash('Nome da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['telefone_referencia']:
            flash('Telefone da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['endereco_referencia']:
            flash('Endereço da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['endereco']:
            flash('Endereço é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['cidade']:
            flash('Cidade é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['estado']:
            flash('Estado é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['cep']:
            flash('CEP é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if dados['cpf_cnpj'] and not validar_cpf_cnpj(dados['cpf_cnpj']):
            flash('CPF/CNPJ inválido.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        if not dados['empresa']:
            flash('Empresa é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
        
        # Atualize os campos do objeto 'cliente' com os dados do formulário
        cliente.nome = dados['nome']
        cliente.cpf_cnpj = dados['cpf_cnpj']
        cliente.rg = dados['rg']
        cliente.email = dados['email']
        cliente.telefone = dados['telefone']
        cliente.telefone_secundario = dados['telefone_secundario']
        cliente.chave_pix = dados['chave_pix']
        cliente.endereco = dados['endereco']
        cliente.cidade = dados['cidade']
        cliente.estado = dados['estado']
        cliente.cep = dados['cep']
        cliente.referencia = dados['referencia']
        cliente.telefone_referencia = dados['telefone_referencia']
        cliente.endereco_referencia = dados['endereco_referencia']
        cliente.observacoes = dados['observacoes']
        cliente.empresa = dados['empresa']
        cliente.atualizado_em = datetime.utcnow()
        
        try:
            # Processar uploads de documentos (se houver)
            files = request.files.getlist('documentos')
            if files:
                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        client_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(cliente_id))
                        os.makedirs(client_upload_folder, exist_ok=True)
                        file_path = os.path.join(client_upload_folder, filename)
                        file.save(file_path)
                        # Manter a query SQL para documentos por enquanto
                        conn = get_db()
                        cur = conn.cursor()
                        cur.execute('INSERT INTO documentos (cliente_id, nome_ficheiro) VALUES (%s, %s)', (cliente_id, filename))
                        conn.commit()
                        cur.close()
                        conn.close()
            
            db.session.commit()
            flash('Cliente atualizado com sucesso!', 'success')
            return redirect(url_for('listar_clientes'))
            
        except IntegrityError:
            db.session.rollback()
            flash('CPF/CNPJ já cadastrado.', 'danger')
            return render_template('cliente_form.html', cliente=cliente)
    
    return render_template('cliente_form.html', cliente=cliente)

@app.route('/cliente/<int:cliente_id>/deletar', methods=['POST'])
@login_required
def deletar_cliente(cliente_id):
    """Deleta um cliente e todas suas cobranças relacionadas."""
    cliente = Customer.query.get(cliente_id)
    
    if not cliente:
        flash('Cliente não encontrado.', 'danger')
        return redirect(url_for('listar_clientes'))
    
    try:
        nome_cliente = cliente.nome  # Pega o nome para a msg flash
        db.session.delete(cliente)
        db.session.commit()
        flash(f'Cliente "{nome_cliente}" e todas as cobranças relacionadas foram excluídos com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir cliente: {str(e)}', 'danger')
    
    return redirect(url_for('listar_clientes'))

@app.route('/cobranca/adicionar', methods=['GET', 'POST'])
@login_required
def adicionar_cobranca():
    """Adiciona uma nova cobrança."""
    clientes = Customer.query.order_by(Customer.nome).all()
    
    if request.method == 'POST':
        # Obter dados do formulário
        cliente_id = int(request.form['cliente_id'])
        descricao = request.form['descricao']
        valor_emprestimo = float(request.form['valor_emprestimo'])
        taxa_juros = float(request.form['taxa_juros'])
        data_primeira_parcela = request.form['data_vencimento']
        
        # Validar se a data não é domingo
        data_venc = datetime.strptime(data_primeira_parcela, '%Y-%m-%d')
        if data_venc.weekday() == 6:  # 6 = domingo (0=segunda, 6=domingo)
            flash('Domingos não são permitidos para data de vencimento. Selecione outro dia.', 'danger')
            return render_template('cobranca_form.html', clientes=clientes, cobranca=None)
        
        # Determinar número de parcelas baseado na taxa de juros
        if taxa_juros == 30:
            numero_parcelas = 10
        elif taxa_juros == 60:
            numero_parcelas = 15
        else:
            flash('Taxa de juros inválida. Use 30% ou 60%.', 'danger')
            return render_template('cobranca_form.html', clientes=clientes, cobranca=None)
        
        # Calcular valor total com juros
        valor_devido_total = valor_emprestimo * (1 + (taxa_juros / 100))
        valor_parcela = valor_devido_total / numero_parcelas
        
        try:
            # 1. Crie a Cobrança principal (o "Pai")
            nova_cobranca = Loan(
                cliente_id=cliente_id,
                descricao=descricao,
                valor_original=valor_emprestimo,
                taxa_juros=taxa_juros,
                valor_total=valor_devido_total,
                data_vencimento=data_venc.date(),
                tipo_cobranca='Parcelada',
                numero_parcelas=numero_parcelas
                # NOTA: O 'client_id' será adicionado depois
            )
            
            # Gerar as parcelas diárias
            data_vencimento_atual = data_venc.date()
            
            for i in range(numero_parcelas):
                # Pular domingos
                while data_vencimento_atual.weekday() == 6:  # 6 = domingo
                    data_vencimento_atual += timedelta(days=1)
                
                # 2. Dentro do loop, crie as Parcelas (as "Filhas")
                nova_parcela = Installment(
                    numero_parcela=i + 1,
                    valor=valor_parcela,
                    data_vencimento=data_vencimento_atual,
                    status='Pendente'
                    # NOTA: O 'client_id' será adicionado depois
                )
                
                # 3. Importante: Associe a parcela à cobrança (Pai -> Filha)
                nova_cobranca.installments.append(nova_parcela)
                
                # Incrementar para a próxima parcela (próximo dia)
                data_vencimento_atual += timedelta(days=1)
            
            # 4. Fora do loop (mas dentro do 'try'), adicione a cobrança principal.
            # O SQLAlchemy salvará todas as parcelas "filhas" em cascata.
            db.session.add(nova_cobranca)
            db.session.commit()
            
            flash(f'Cobrança criada com sucesso! {numero_parcelas} parcelas diárias foram geradas.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar cobrança: {str(e)}', 'danger')
        
        return redirect(url_for('index'))
    
    return render_template('cobranca_form.html', clientes=clientes, cobranca=None)

@app.route('/cobranca/<int:cobranca_id>/cancelar', methods=['POST'])
@login_required
def cancelar_cobranca(cobranca_id):
    """DELETA uma cobrança e todos os seus registros associados (parcelas, pagamentos)."""
    # Buscar a cobrança ou retornar 404
    cobranca = Loan.query.get_or_404(cobranca_id)
    
    # Armazena o ID do cliente para o redirecionamento
    cliente_id_redirect = cobranca.cliente_id
    
    try:
        # Deletar a cobrança. 
        # O 'cascade="all, delete-orphan"' no modelo
        # cuidará das parcelas e pagamentos.
        db.session.delete(cobranca)
        db.session.commit()
        flash('Cobrança e todos os seus dados foram DELETADOS com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar cobrança: {str(e)}', 'danger')
    
    # Redireciona de volta para a página do cliente
    return redirect(url_for('visualizar_cliente', cliente_id=cliente_id_redirect))

@app.route('/cobranca/<int:id>/registrar_pagamento', methods=['POST'])
@login_required
def registrar_pagamento(id):
    """Registra um pagamento genérico para uma cobrança."""
    # Busca a cobrança usando ORM
    cobranca = Loan.query.get_or_404(id)
    
    try:
        # Obter dados do formulário
        valor_a_pagar = float(request.form['valor_pago'])
        observacao_pagamento = request.form.get('observacao_pagamento', '')
        
        if valor_a_pagar <= 0:
            flash('O valor do pagamento deve ser maior que zero.', 'warning')
            return redirect(url_for('visualizar_cliente', cliente_id=cobranca.cliente_id))
        
        # 1. Criar o registro de pagamento
        novo_pagamento = Payment(
            cobranca_id=cobranca.id,
            valor_pago=valor_a_pagar,
            observacao=observacao_pagamento,
            forma_pagamento='Dinheiro',
            usuario_id=session.get('usuario_id'),
            client_id=cobranca.cliente_id
        )
        
        # 2. Atualizar o valor pago na cobrança principal
        # Convertemos para float para garantir a soma correta
        valor_atual = float(cobranca.valor_pago or 0)
        cobranca.valor_pago = valor_atual + valor_a_pagar
        cobranca.atualizado_em = datetime.utcnow()
        
        # 3. Criar registro no histórico (para manter compatibilidade com a visualização)
        novo_historico = PaymentHistory(
            cobranca_id=cobranca.id,
            cliente_id=cobranca.cliente_id,
            valor_pago=valor_a_pagar,
            forma_pagamento='Dinheiro',
            observacoes=observacao_pagamento,
            usuario_id=session.get('usuario_id'),
            client_id=cobranca.cliente_id
        )
        
        # Adiciona tudo à sessão e salva
        db.session.add(novo_pagamento)
        db.session.add(novo_historico)
        db.session.commit()
        
        flash(f'Pagamento de R$ {valor_a_pagar:.2f} registrado com sucesso.', 'info')
        
    except ValueError:
        flash('Valor inválido informado.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar pagamento: {str(e)}', 'danger')

    return redirect(url_for('visualizar_cliente', cliente_id=cobranca.cliente_id))


@app.route('/cobranca/<int:cobranca_id>/pagamentos')
@login_required
def visualizar_pagamentos_cobranca(cobranca_id):
    """Visualiza o histórico de pagamentos de uma cobrança específica."""
    # Busca a cobrança (e o cliente via relacionamento)
    cobranca = Loan.query.get_or_404(cobranca_id)
    
    # Busca pagamentos ordenados por data (mais recente primeiro)
    pagamentos = Payment.query.filter_by(cobranca_id=cobranca_id)\
        .order_by(Payment.data_pagamento.desc(), Payment.id.desc()).all()
    
    return render_template('pagamentos_cobranca.html', 
                         cobranca=cobranca, 
                         pagamentos=pagamentos)

@app.route('/parcela/<int:id>/pagar', methods=['POST'])
@login_required
def marcar_parcela_paga(id):
    """Marca uma parcela como paga e atualiza o valor_pago da cobrança."""
    # Busca a parcela
    parcela = Installment.query.get_or_404(id)
    cobranca = parcela.loan # Pega a cobrança pai automaticamente
    
    if parcela.status == 'Pago':
        flash('Esta parcela já foi paga.', 'info')
        return redirect(url_for('visualizar_cliente', cliente_id=cobranca.cliente_id))
    
    try:
        # 1. Atualizar a Parcela
        parcela.status = 'Pago'
        parcela.valor_pago = parcela.valor
        parcela.forma_pagamento = 'Dinheiro'
        parcela.data_pagamento = date.today()
        parcela.atualizado_em = datetime.utcnow()
        
        # 2. Atualizar o valor pago na cobrança principal
        valor_atual_cobranca = float(cobranca.valor_pago or 0)
        valor_parcela = float(parcela.valor)
        cobranca.valor_pago = valor_atual_cobranca + valor_parcela
        cobranca.atualizado_em = datetime.utcnow()
        
        # 3. Registrar no histórico de pagamentos (para auditoria)
        novo_historico = PaymentHistory(
            cobranca_id=cobranca.id,
            cliente_id=cobranca.cliente_id,
            valor_pago=parcela.valor,
            forma_pagamento='Dinheiro',
            observacoes=f'Pagamento da parcela {parcela.numero_parcela}',
            usuario_id=session.get('usuario_id'),
            client_id=cobranca.cliente_id
        )
        db.session.add(novo_historico)
        
        # 4. Verificar se a cobrança foi totalmente quitada
        # Se o valor pago >= valor total, encerra a cobrança
        if float(cobranca.valor_pago) >= float(cobranca.valor_total):
            cobranca.status = 'Pago'
            cobranca.data_pagamento = date.today()
            flash(f'Parcela {parcela.numero_parcela} paga. Todas as parcelas foram pagas e a cobrança foi liquidada!', 'success')
        else:
            flash(f'Parcela {parcela.numero_parcela} marcada como paga.', 'success')
        
        # Salva todas as alterações (Parcela, Cobrança e Histórico)
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao processar pagamento: {str(e)}', 'danger')
    
    return redirect(url_for('visualizar_cliente', cliente_id=cobranca.cliente_id))

@app.route('/parcela/<int:id>/editar_multa', methods=['POST'])
@login_required
def editar_multa_parcela(id):
    """Edita a multa manual de uma parcela."""
    # Busca a parcela
    parcela = Installment.query.get_or_404(id)
    
    multa_input = request.form.get('multa_manual_parcela')

    try:
        if multa_input and multa_input.strip() != '':
            multa_valor = float(multa_input)
            if multa_valor < 0:
                flash('O valor da multa manual não pode ser negativo.', 'danger')
            else:
                parcela.multa_manual = multa_valor
                parcela.atualizado_em = datetime.utcnow()
                db.session.commit()
                flash(f'Multa manual de R$ {multa_valor:.2f} definida para a parcela {parcela.numero_parcela}.', 'success')
        else: 
            # Remove a multa se o campo estiver vazio
            parcela.multa_manual = None
            parcela.atualizado_em = datetime.utcnow()
            db.session.commit()
            flash(f'Multa manual removida da parcela {parcela.numero_parcela}.', 'info')

    except ValueError:
        flash('Valor da multa inválido. Por favor, insira um número.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar multa: {str(e)}', 'danger')

    return redirect(url_for('visualizar_cliente', cliente_id=parcela.client_id))

@app.route('/parcela/<int:id>/editar_data', methods=['POST'])
@login_required
def editar_data_parcela(id):
    """Edita a data de vencimento de uma parcela."""
    # Busca a parcela
    parcela = Installment.query.get_or_404(id)
    
    nova_data_str = request.form.get('nova_data_vencimento')
    
    if not nova_data_str:
        flash('Data de vencimento inválida.', 'danger')
        return redirect(url_for('visualizar_cliente', cliente_id=parcela.client_id))
    
    try:
        nova_data = datetime.strptime(nova_data_str, '%Y-%m-%d').date()
        
        # Validação para não permitir data no domingo
        if nova_data.weekday() == 6:  # 6 = Domingo
            flash('A data de vencimento não pode ser num domingo.', 'warning')
            return redirect(url_for('visualizar_cliente', cliente_id=parcela.client_id))
        
        # Atualiza a data
        parcela.data_vencimento = nova_data
        parcela.atualizado_em = datetime.utcnow()
        db.session.commit()
        
        flash(f'Data de vencimento da parcela {parcela.numero_parcela} atualizada para {nova_data.strftime("%d/%m/%Y")}.', 'success')
    except ValueError:
        flash('Formato de data inválido.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar data: {str(e)}', 'danger')
    
    return redirect(url_for('visualizar_cliente', cliente_id=parcela.client_id))

@app.route('/cobrancas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cobranca(id):
    """Edita uma cobrança existente."""
    cobranca = Loan.query.get_or_404(id)
    cliente = cobranca.customer

    if not cliente:
        flash('Cliente não encontrado.', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            valor_emprestimo_str = request.form.get('valor_emprestimo')
            if not valor_emprestimo_str:
                flash('Valor emprestado é obrigatório.', 'danger')
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            novo_valor_emprestimo = float(valor_emprestimo_str)
            if novo_valor_emprestimo <= 0:
                flash('Valor emprestado deve ser maior que zero.', 'danger')
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            nova_data_vencimento_str = request.form.get('data_vencimento')
            if not nova_data_vencimento_str:
                flash('Data de vencimento é obrigatória.', 'danger')
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            nova_data_vencimento = datetime.strptime(nova_data_vencimento_str, '%Y-%m-%d').date()
            
            if nova_data_vencimento.weekday() == 6:
                flash('Domingos não são permitidos para data de vencimento. Selecione outro dia.', 'danger')
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            taxa_juros_str = request.form.get('taxa_juros')
            if taxa_juros_str:
                taxa_juros = float(taxa_juros_str)
            else:
                taxa_juros = float(cobranca.taxa_juros)
            
            # O ERRO ESTAVA AQUI (ALINHAMENTO)
            if taxa_juros == 30:
                numero_parcelas = 10
            elif taxa_juros == 60:
                numero_parcelas = 15
            else:
                numero_parcelas = int(request.form.get('numero_parcelas', cobranca.numero_parcelas))

            valor_total = novo_valor_emprestimo * (1 + (taxa_juros / 100))
            valor_parcela = valor_total / numero_parcelas
            
            cobranca.installments = []
            
            PaymentHistory.query.filter_by(cobranca_id=id).delete()
            
            data_vencimento_atual = nova_data_vencimento
            
            for i in range(numero_parcelas):
                while data_vencimento_atual.weekday() == 6:
                    data_vencimento_atual += timedelta(days=1)
                
                nova_parcela = Installment(
                    numero_parcela=i + 1,
                    valor=valor_parcela,
                    data_vencimento=data_vencimento_atual,
                    status='Pendente'
                )
                cobranca.installments.append(nova_parcela)
                data_vencimento_atual += timedelta(days=1)
            
            cobranca.valor_original = novo_valor_emprestimo
            cobranca.valor_total = valor_total
            cobranca.taxa_juros = taxa_juros
            cobranca.data_vencimento = nova_data_vencimento
            cobranca.numero_parcelas = numero_parcelas
            cobranca.valor_pago = 0
            cobranca.status = 'Pendente'
            cobranca.atualizado_em = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'Cobrança atualizada com sucesso! {numero_parcelas} parcelas foram re-geradas.', 'success')
            return redirect(url_for('visualizar_cliente', cliente_id=cobranca.cliente_id))
            
        except ValueError as e:
            db.session.rollback()
            flash(f'Erro nos valores: {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar cobrança: {str(e)}', 'danger')
    
    return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
    return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)




# --- Rotas de Usuários ---
@app.route('/usuarios')
@login_required
@adm_required
def listar_usuarios():
    """Lista todos os usuários do sistema."""
    usuarios = User.query.order_by(User.nome).all()
    
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/usuario/adicionar', methods=['GET', 'POST'])
@login_required
@adm_required
def adicionar_usuario():
    """Adiciona um novo usuário."""
    if request.method == 'POST':
        dados = {
            'nome': request.form['nome'],
            'email': request.form['email'],
            'senha': request.form['senha'],
            'nivel': request.form.get('nivel', 'Operador')
        }
        
        # Automatically derive tipo from nivel
        dados['tipo'] = 'admin' if dados['nivel'] == 'ADM' else 'operador'
        
        try:
            senha_hash = generate_password_hash(dados['senha'])
            novo_usuario = User(
                nome=dados['nome'],
                email=dados['email'],
                senha=senha_hash,
                tipo=dados['tipo'],
                nivel=dados['nivel']
            )
            db.session.add(novo_usuario)
            db.session.commit()
            flash('Usuário adicionado com sucesso!', 'success')
            return redirect(url_for('listar_usuarios'))
        except IntegrityError:
            db.session.rollback()
            flash('Email já cadastrado.', 'danger')
    
    return render_template('usuario_form.html', usuario=None)

@app.route('/usuario/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
@adm_required
def editar_usuario(usuario_id):
    """Edita um usuário existente."""
    usuario = User.query.get(usuario_id)
    if not usuario:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('listar_usuarios'))
    
    if request.method == 'POST':
        dados = {
            'nome': request.form['nome'],
            'email': request.form['email'],
            'nivel': request.form.get('nivel', 'Operador')
        }
        
        # Automatically derive tipo from nivel
        dados['tipo'] = 'admin' if dados['nivel'] == 'ADM' else 'operador'
        
        try:
            # Atualizar campos do usuário
            usuario.nome = dados['nome']
            usuario.email = dados['email']
            usuario.tipo = dados['tipo']
            usuario.nivel = dados['nivel']
            
            # Se uma nova senha foi fornecida
            if request.form.get('senha'):
                usuario.senha = generate_password_hash(request.form['senha'])
            
            db.session.commit()
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('listar_usuarios'))
        except IntegrityError:
            db.session.rollback()
            flash('Email já cadastrado.', 'danger')
    
    return render_template('usuario_form.html', usuario=usuario)

@app.route('/usuario/<int:usuario_id>/deletar', methods=['POST'])
@login_required
@adm_required
def excluir_usuario(usuario_id):
    """Exclui um usuário do sistema."""
    usuario = User.query.get(usuario_id)
    if not usuario:
        flash('Usuário não encontrado.', 'danger')
        return redirect(url_for('listar_usuarios'))
    
    try:
        nome_usuario = usuario.nome
        db.session.delete(usuario)
        db.session.commit()
        flash(f'Usuário "{nome_usuario}" foi excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir usuário: {str(e)}', 'danger')
    
    return redirect(url_for('listar_usuarios'))

# --- Rotas do Calendário ---
@app.route('/calendario')
@login_required
def calendario():
    """Renderiza a página do calendário."""
    return render_template('calendario.html')

@app.route('/api/eventos')
@login_required
def api_eventos():
    """API para buscar eventos do calendário."""
    # Busca cobranças que têm data de vencimento definida
    cobrancas = Loan.query.filter(Loan.data_vencimento.isnot(None)).all()
    
    eventos = []
    for cobranca in cobrancas:
        # Verifica se o cliente existe (segurança)
        if cobranca.customer:
            eventos.append({
                'title': f"{cobranca.customer.nome} - R$ {float(cobranca.valor_original):.2f}",
                'start': cobranca.data_vencimento.isoformat(),
                'url': url_for('visualizar_cliente', cliente_id=cobranca.cliente_id)
            })
    
    return jsonify(eventos)

# --- Rota para servir uploads ---
@app.route('/uploads/<int:cliente_id>/<filename>')
@login_required
def uploaded_file(cliente_id, filename):
    # --- CORREÇÃO ---
    # 1. Obter o caminho absoluto para a pasta 'uploads' usando app.root_path
    # app.root_path é o caminho absoluto para o diretório onde main.py está
    uploads_dir = os.path.join(app.root_path, app.config['UPLOAD_FOLDER'])
    
    # 2. Construir o caminho para a pasta específica do cliente
    client_dir = os.path.join(uploads_dir, str(cliente_id))
    
    # 3. Construir o caminho completo para o arquivo
    file_path = os.path.join(client_dir, filename)
    
    # DEBUG: Imprime o caminho completo
    print(f"--- DEBUG: A tentar servir o ficheiro: {file_path} ---")
    
    # 4. Verificar se o arquivo existe
    if not os.path.isfile(file_path):
        print(f"--- DEBUG ERROR: Ficheiro não encontrado: {file_path} ---")
        abort(404)
    
    # 5. Tenta servir o ficheiro
    try:
        # Passamos o diretório absoluto do cliente e o nome do arquivo
        print(f"--- DEBUG SUCCESS: A servir {filename} de {client_dir} ---")
        return send_from_directory(client_dir, filename, as_attachment=False)
    except Exception as e:
        print(f"--- DEBUG ERROR: Erro ao servir ficheiro: {e} ---")
        abort(500)





# --- Rotas de Relatórios ---
@app.route('/api/relatorios/kpis')
@login_required
@gerente_required
def api_relatorios_kpis():
    """API para buscar KPIs dos relatórios."""
    
    # 1. Counts básicos
    total_clientes = Customer.query.count()
    cobrancas_pendentes = Loan.query.filter_by(status='Pendente').count()
    cobrancas_pagas = Loan.query.filter_by(status='Pago').count()
    
    # Cobranças Vencidas
    cobrancas_vencidas = Loan.query.filter(
        Loan.status == 'Pendente', 
        Loan.data_vencimento < date.today()
    ).count()
    
    # 2. Saldo Devedor Total (Soma das parcelas pendentes + multas)
    # Usamos func.sum e func.coalesce para tratar valores nulos
    saldo_devedor_total = db.session.query(
        func.sum(Installment.valor + func.coalesce(Installment.multa_manual, 0))
    ).join(Loan).filter(
        Loan.status == 'Pendente',
        Installment.status == 'Pendente'
    ).scalar() or 0
    
    # 3. Total Recebido no Mês
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    
    total_recebido_mes = db.session.query(func.sum(Payment.valor_pago))\
        .filter(Payment.data_pagamento >= primeiro_dia_mes).scalar() or 0
    
    return jsonify({
        'total_clientes': total_clientes,
        'cobrancas_pendentes': cobrancas_pendentes,
        'cobrancas_vencidas': cobrancas_vencidas,
        'cobrancas_pagas': cobrancas_pagas,
        'saldo_devedor_total': float(saldo_devedor_total),
        'total_recebido_mes': float(total_recebido_mes),
    })

@app.route('/relatorios')
@login_required
@gerente_required
def relatorios():
    """Página principal de relatórios."""
    return render_template('relatorios.html')




@app.route('/relatorios/clientes', methods=['GET', 'POST'])
@login_required
@gerente_required
def gerar_relatorio_clientes():
    """Gera relatório de clientes em Excel."""
    try:
        # Buscar dados via ORM
        clientes = Customer.query.order_by(Customer.nome).all()
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Clientes"

        # --- Estilo do Cabeçalho ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        headers = ['ID', 'Nome', 'CPF/CNPJ', 'Telefone', 'Email', 'Cidade', 'Data de Cadastro']
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # --- Escreve os Dados ---
        for cliente in clientes:
            ws.append([
                cliente.id,
                cliente.nome,
                cliente.cpf_cnpj,
                cliente.telefone,
                cliente.email or 'N/A',
                cliente.cidade,
                cliente.criado_em.strftime('%d/%m/%Y') if cliente.criado_em else 'N/A'
            ])

        # Autoajuste
        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].auto_size = True

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=relatorio_clientes.xlsx"}
        )
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('relatorios'))

@app.route('/relatorios/cobrancas', methods=['GET', 'POST'])
@login_required
@gerente_required
def gerar_relatorio_cobrancas():
    """Gera relatório de cobranças em Excel."""
    try:
        # Buscar cobranças e carregar o cliente junto (joinedload) para ser rápido
        cobrancas = Loan.query.options(db.joinedload(Loan.customer))\
            .order_by(Loan.data_vencimento.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Cobranças"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")

        headers = [
            'ID Cobrança', 'Nome Cliente', 'Valor Original', 'Valor Devido (c/ juros)', 
            'Total Pago', 'Nº de Pagamentos', 'Data Vencimento', 'Status Pagamento'
        ]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for cobranca in cobrancas:
            total_pago = float(cobranca.valor_pago or 0)
            # Conta quantos pagamentos existem para essa cobrança
            numero_pagamentos = cobranca.payments.count()

            ws.append([
                cobranca.id, 
                cobranca.customer.nome, 
                cobranca.valor_original,
                cobranca.valor_total or cobranca.valor_original,
                total_pago,
                numero_pagamentos,
                cobranca.data_vencimento.strftime('%d/%m/%Y') if cobranca.data_vencimento else 'N/A',
                'Pago' if cobranca.status == 'Pago' else 'Pendente'
            ])

        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].auto_size = True

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=relatorio_cobrancas.xlsx"}
        )
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('relatorios'))

# --- Funções para Templates ---
@app.context_processor
def utility_processor():
    def get_user_nivel():
        """Retorna o nível do usuário logado."""
        return session.get('usuario_nivel', 'Operador')
    
    def can_access_reports():
        """Verifica se o usuário pode acessar relatórios."""
        return session.get('usuario_nivel') in ['Gerente', 'ADM']
    
    def can_access_admin():
        """Verifica se o usuário pode acessar funcionalidades administrativas."""
        return session.get('usuario_nivel') == 'ADM'
    
    return dict(
        get_user_nivel=get_user_nivel,
        can_access_reports=can_access_reports,
        can_access_admin=can_access_admin
    )

# --- Execução da Aplicação ---
if __name__ == '__main__':
    # Verificar se DATABASE_URL está configurada
    if not DATABASE_URL:
        print("❌ ERRO: Variável de ambiente DATABASE_URL não está configurada!")
        print("Configure a DATABASE_URL antes de executar o app.")
        print("Exemplo: export DATABASE_URL='postgresql://usuario:senha@localhost:5432/crm_db'")
        sys.exit(1)

    print("🚀 Iniciando aplicação Flask...")
    port = int(os.environ.get("PORT", 5000))  # Pega a porta do ambiente ou usa 5000 como padrão
    app.run(host='0.0.0.0', port=port, debug=True)

